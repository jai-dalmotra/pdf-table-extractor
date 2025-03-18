pip install pymupdf pdfplumber pandas openpyxl

import fitz  # PyMuPDF
import pdfplumber
import pandas as pd
from openpyxl import Workbook
import unicodedata
import re
import numpy as np
import os

# Helper function to clean and normalize text
def clean_text(text):
    text = unicodedata.normalize("NFKD", text)
    return ''.join(char for char in text if char.isprintable() and ord(char) >= 32)

# Sanitize text to remove illegal characters
def sanitize_for_excel(text):
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]')
    return ILLEGAL_CHARACTERS_RE.sub("", str(text))

# Extracts and clusters words by their spatial location
def extract_spatial_tables_from_pdf(pdf_path, workbook):
    doc = fitz.open(pdf_path)
    sheet = workbook.create_sheet(title="Spatial Data")

    for page_number, page in enumerate(doc):
        print(f"Processing page {page_number + 1}")
        words = page.get_text("words")

        if not words:
            continue

        rows = []

        for x0, y0, x1, y1, text, *_ in words:
            text = sanitize_for_excel(clean_text(text))
            rows.append((x0, y0, text))

        rows.sort(key=lambda w: (w[1], w[0]))

        clustered_rows = []
        current_row = []
        current_y = None

        for x, y, text in rows:
            if current_y is None or abs(y - current_y) > 2:  # Threshold for row change
                if current_row:
                    clustered_rows.append(current_row)
                current_row = []
                current_y = y

            current_row.append((x, text))

        if current_row:
            clustered_rows.append(current_row)

        for row in clustered_rows:
            row.sort(key=lambda w: w[0])
            row_texts = [text for _, text in row]

            try:
                sheet.append(row_texts)
            except Exception as e:
                print(f"Error writing row: {row_texts}. Error: {e}")

# Extract text-based tables using pdfplumber
def extract_text_based_tables(pdf_path, workbook):
    with pdfplumber.open(pdf_path) as pdf:
        sheet = workbook.create_sheet(title="Text Data")

        for page in pdf.pages:
            words = page.extract_words()
            if not words:
                continue

            lines = {}
            for word in words:
                y = round(word["top"])
                if y not in lines:
                    lines[y] = []
                lines[y].append(word["text"])

            current_headers = {}
            table_data = {}

            for line in lines.values():
                if ":" in line:
                    key_index = line.index(":") - 1 if line.index(":") > 0 else 0
                    key = " ".join(line[:key_index + 1]).strip()
                    value = " ".join(line[key_index + 2:]).strip()

                    if key and key not in current_headers:
                        current_headers[key] = len(current_headers)
                        table_data[key] = []

                    table_data[key].append(value)
                else:
                    if table_data:
                        for key in table_data:
                            table_data[key][-1] += " " + " ".join(line)

            if table_data:
                headers = list(table_data.keys())
                sheet.append(headers)

                max_len = max(len(col) for col in table_data.values())

                for i in range(max_len):
                    row = [table_data.get(header, [""])[i] if i < len(table_data.get(header, [])) else "" for header in headers]
                    sheet.append(row)

def extract_tables_from_pdf(pdf_path, excel_path):
    workbook = Workbook()

    extract_spatial_tables_from_pdf(pdf_path, workbook)
    extract_text_based_tables(pdf_path, workbook)

    workbook.remove(workbook["Sheet"]) 

    workbook.save(excel_path)
    print(f"Tables extracted and saved to: {excel_path}")

if __name__ == "__main__":
    pdf_path = "/kaggle/input/socreme-data/test6.pdf" 
    excel_path = "output.xlsx"

    if os.path.exists(pdf_path):
        extract_tables_from_pdf(pdf_path, excel_path)
    else:
        print("Error: PDF file not found.")
